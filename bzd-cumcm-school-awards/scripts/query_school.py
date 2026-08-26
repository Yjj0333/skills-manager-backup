#!/usr/bin/env python3
import argparse,json,re
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook
def n(v):return re.sub(r"\s+","",str(v or "").strip().lower().replace("（","(").replace("）",")"))
def s(a,b):
 a,b=n(a),n(b)
 return 1 if a==b else (.9 if a and(a in b or b in a) else SequenceMatcher(None,a,b).ratio())
def main():
 p=argparse.ArgumentParser();p.add_argument("--school",required=True);p.add_argument("--region",default="");a=p.parse_args();f=Path(__file__).resolve().parents[1]/"assets"/"高教社杯国赛学校综合统计与2026预测.xlsx";w=load_workbook(f,read_only=True,data_only=True)["学校综合统计"];h=[c.value for c in next(w.iter_rows())];rs=[dict(zip(h,r))for r in w.iter_rows(min_row=2,values_only=True)];z=sorted([(s(a.school,r["学校名称"]),not a.region or n(a.region)==n(r["所属赛区"]),r)for r in rs],key=lambda x:(x[1],x[0]),reverse=True);q=[x for x in z if x[1]and x[0]>=.72]
 if q and(len(q)==1 or q[0][0]-q[1][0]>=.08):
  r=q[0][2];o={"status":"ok","学校名称":r["学校名称"],"所属赛区":r["所属赛区"],"年度记录":{str(y):{"一等奖":r[f"{y}一等奖"],"二等奖":r[f"{y}二等奖"],"合计":r[f"{y}国奖合计"]}for y in range(2021,2026)},"五年汇总":{"一等奖":r["五年一等奖合计"],"二等奖":r["五年二等奖合计"],"合计":r["五年国奖合计"]},"2026国奖预测":r["2026国奖预测"],"最高频指导老师":r["最高频指导老师"],"出现次数":r["出现次数"]}
 elif q:o={"status":"ambiguous","candidates":[{"学校名称":x[2]["学校名称"],"所属赛区":x[2]["所属赛区"]}for x in q[:8]]}
 else:o={"status":"not_found","年度记录":{str(y):{"一等奖":0,"二等奖":0,"合计":0}for y in range(2021,2026)},"五年汇总":{"一等奖":0,"二等奖":0,"合计":0},"2026国奖经验概率":"6.81%","message":"非常遗憾，按当前榜单口径，您所在的高校过去5年没有获得国奖成绩。对于这类学校，本年度可能获得国奖的经验概率为6.81%。该数字不是个人获奖概率或官方结论。"}
 print(json.dumps(o,ensure_ascii=False,indent=2))
if __name__=="__main__":main()
